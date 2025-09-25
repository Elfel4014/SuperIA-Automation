from openpyxl import Workbook, load_workbook
from openpyxl.styles import Font, Border, Side

class ExcelAutomator:
    def __init__(self):
        pass

    def create_workbook(self, filename="planilha_superia.xlsx", sheet_name="Dados SuperIA"):
        try:
            wb = Workbook()
            ws = wb.active
            ws.title = sheet_name
            wb.save(filename)
            print(f"Pasta de trabalho Excel \'{filename}\' criada com sucesso.")
            return True
        except Exception as e:
            print(f"Erro ao criar pasta de trabalho Excel: {e}")
            return False

    def write_data_to_sheet(self, filename, sheet_name, data, start_row=1, start_col=1, include_headers=True):
        try:
            # Tenta carregar a pasta de trabalho existente, se não existir, cria uma nova
            try:
                wb = load_workbook(filename)
            except FileNotFoundError:
                wb = Workbook()
                ws = wb.active
                ws.title = sheet_name
                wb.save(filename) # Salva a nova pasta de trabalho para que possa ser carregada
                wb = load_workbook(filename)

            if sheet_name not in wb.sheetnames:
                ws = wb.create_sheet(sheet_name)
            else:
                ws = wb[sheet_name]

            row_offset = 0
            if include_headers and data and isinstance(data[0], dict):
                headers = list(data[0].keys())
                for col_idx, header in enumerate(headers, start=start_col):
                    ws.cell(row=start_row, column=col_idx, value=header).font = Font(bold=True)
                row_offset = 1
                # Escreve os dados
                for row_idx, row_data in enumerate(data, start=start_row + row_offset):
                    for col_idx, header in enumerate(headers, start=start_col):
                        ws.cell(row=row_idx, column=col_idx, value=row_data[header])
            elif data and isinstance(data[0], (list, tuple)):
                for row_idx, row_data in enumerate(data, start=start_row):
                    for col_idx, cell_value in enumerate(row_data, start=start_col):
                        ws.cell(row=row_idx, column=col_idx, value=cell_value)
            else:
                print("Formato de dados não suportado. Use lista de listas/tuplas ou lista de dicionários.")
                return False

            wb.save(filename)
            print(f"Dados escritos na planilha \'{sheet_name}\' do arquivo \'{filename}\' com sucesso.")
            return True
        except Exception as e:
            print(f"Erro ao escrever dados na planilha Excel: {e}")
            return False

    def read_data_from_sheet(self, filename, sheet_name, min_row=1, max_row=None, min_col=1, max_col=None):
        try:
            wb = load_workbook(filename)
            ws = wb[sheet_name]

            data = []
            for row in ws.iter_rows(min_row=min_row, max_row=max_row, min_col=min_col, max_col=max_col):
                row_data = [cell.value for cell in row]
                data.append(row_data)
            print(f"Dados lidos da planilha \'{sheet_name}\' do arquivo \'{filename}\' com sucesso.")
            return data
        except FileNotFoundError:
            print(f"Arquivo \'{filename}\' não encontrado.")
            return None
        except KeyError:
            print(f"Planilha \'{sheet_name}\' não encontrada no arquivo \'{filename}\'")
            return None
        except Exception as e:
            print(f"Erro ao ler dados da planilha Excel: {e}")
            return None

    def add_formula(self, filename, sheet_name, cell_address, formula):
        try:
            wb = load_workbook(filename)
            ws = wb[sheet_name]
            ws[cell_address] = formula
            wb.save(filename)
            print(f"Fórmula \'{formula}\' adicionada à célula \'{cell_address}\' na planilha \'{sheet_name}\' com sucesso.")
            return True
        except Exception as e:
            print(f"Erro ao adicionar fórmula: {e}")
            return False

if __name__ == "__main__":
    automator = ExcelAutomator()
    test_excel = "test_superia_excel.xlsx"
    test_sheet = "Relatório"

    print("\n--- Criando pasta de trabalho Excel ---")
    automator.create_workbook(test_excel, test_sheet)

    print("\n--- Escrevendo dados com cabeçalhos ---")
    data_with_headers = [
        {"Produto": "Teclado", "Quantidade": 5, "PrecoUnitario": 150.00},
        {"Produto": "Mouse", "Quantidade": 10, "PrecoUnitario": 75.00},
        {"Produto": "Monitor", "Quantidade": 2, "PrecoUnitario": 1200.00}
    ]
    automator.write_data_to_sheet(test_excel, test_sheet, data_with_headers, include_headers=True)

    print("\n--- Adicionando fórmulas ---")
    automator.add_formula(test_excel, test_sheet, "D1", "Total")
    automator.add_formula(test_excel, test_sheet, "D2", "=B2*C2")
    automator.add_formula(test_excel, test_sheet, "D3", "=B3*C3")
    automator.add_formula(test_excel, test_sheet, "D4", "=B4*C4")
    automator.add_formula(test_excel, test_sheet, "D5", "=SUM(D2:D4)")

    print("\n--- Lendo dados da planilha ---")
    read_data = automator.read_data_from_sheet(test_excel, test_sheet)
    if read_data:
        for row in read_data:
            print(row)

    print(f"\nVerifique o arquivo \'{test_excel}\' para ver o resultado.")

